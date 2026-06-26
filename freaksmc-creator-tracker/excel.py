"""Excel (.xlsx) import/export for the creator tracker.

Export writes a formatted workbook with a Creators sheet (one row per partner,
including the computed commission) plus a small Summary sheet. Import reads a
sheet back by matching column headers to the canonical field names, so a file
produced by export round-trips cleanly and hand-built sheets work too.
"""

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db

# Human-friendly column headers, in export order. The computed columns
# (Effective Rate / Commission Owed) are export-only and ignored on import.
EXPORT_COLUMNS = [
    ("name", "Creator Name"),
    ("youtube_link", "YouTube Channel"),
    ("subscriber_count", "Subscribers"),
    ("niche", "Niche"),
    ("discord_link", "Discord Server"),
    ("discord_username", "Discord Username"),
    ("contact_date", "Contact Date"),
    ("status", "Ticket Status"),
    ("affiliate_code", "Affiliate Code"),
    ("signups", "Signups"),
    ("rate", "Per-Signup Rate"),
    ("effective_rate", "Effective Rate"),
    ("commission", "Commission Owed"),
    ("notes", "Notes"),
]

# Map a normalised header string -> canonical field name, for import. We accept
# both the friendly headers above and the raw field names.
def _header_lookup():
    lookup = {}
    for field, header in EXPORT_COLUMNS:
        lookup[_norm(header)] = field
        lookup[_norm(field)] = field
    return lookup


def _norm(text):
    return "".join(ch for ch in str(text).lower() if ch.isalnum())


HEADER_FILL = PatternFill("solid", fgColor="2F3640")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_to_bytes():
    """Build the workbook in memory and return it as bytes for download."""
    records = db.list_creators(sort="contact_date", direction="desc")
    summary = db.compute_summary(records)
    settings = db.load_settings()
    currency = settings.get("currency", "$")

    wb = Workbook()

    # --- Creators sheet ----------------------------------------------------
    ws = wb.active
    ws.title = "Creators"
    headers = [h for _, h in EXPORT_COLUMNS]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        row = []
        for field, _ in EXPORT_COLUMNS:
            value = rec.get(field)
            if field in ("rate", "effective_rate", "commission") and value is not None:
                value = round(float(value), 2)
            row.append(value)
        ws.append(row)

    # Auto-ish column widths based on the longest cell in each column.
    for col_idx, (_field, header) in enumerate(EXPORT_COLUMNS, start=1):
        longest = len(header)
        for rec in records:
            longest = max(longest, len(str(rec.get(_field, "") or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + 3, 45)
    ws.freeze_panes = "A2"

    # --- Summary sheet -----------------------------------------------------
    s = wb.create_sheet("Summary")
    s.append(["Metric", "Value"])
    for col_idx in (1, 2):
        cell = s.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    rows = [
        ("Total outreached", summary["total"]),
        ("Responded", summary["responded"]),
        ("Live partnerships", summary["live"]),
        ("Response rate", f"{summary['response_rate']:.1f}%"),
        ("Live conversion rate", f"{summary['conversion_rate']:.1f}%"),
        ("Total signups", summary["total_signups"]),
        ("Total commission owed", f"{currency}{summary['total_commission']:.2f}"),
    ]
    for label, value in rows:
        s.append([label, value])
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def import_from_filestream(stream):
    """Read creator rows from an uploaded .xlsx stream.

    Returns (imported_count, skipped_count). Rows without a creator name are
    skipped. Unknown columns are ignored; missing columns fall back to defaults.
    """
    wb = load_workbook(stream, data_only=True)
    ws = wb["Creators"] if "Creators" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return 0, 0

    lookup = _header_lookup()
    # Build a column-index -> field map from the header row.
    col_to_field = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        field = lookup.get(_norm(header))
        if field:
            col_to_field[idx] = field

    imported = 0
    skipped = 0
    for raw in rows:
        if raw is None:
            continue
        data = {}
        for idx, field in col_to_field.items():
            if idx < len(raw):
                data[field] = raw[idx]
        name = str(data.get("name") or "").strip()
        if not name:
            skipped += 1
            continue
        db.create_creator(data)
        imported += 1
    return imported, skipped
